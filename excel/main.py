from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('Fuck')
ws2 = wb.create_sheet('Fuck', 0)
ws3 = wb.create_sheet('Fuck', -1)
ws.title = 'New Title'



